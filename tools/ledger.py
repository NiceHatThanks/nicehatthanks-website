"""Local Excel finance ledger for NiceHatThanks orders.

The ledger is intentionally stored outside the git repository because it can
contain customer/order information and manual financial data.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


if os.name == "nt" and os.environ.get("LOCALAPPDATA"):
    DATA_DIR = Path(os.environ["LOCALAPPDATA"]) / "NiceHatThanks" / "finance"
else:
    DATA_DIR = Path.home() / ".nicehatthanks" / "finance"

PRODUCT_DEFAULTS = {
    "scoopy": ("Scoopy", 15.99),
    "scoopy_compact": ("Scoopy Compact", 12.49),
    "pcba_mmwave": ("Populated PCBA + mmWave", 11.49),
    "pcba": ("Populated PCBA", 8.49),
}

DARK = "1C211B"
CREAM = "F5F0E6"
WHITE = "FFFDF8"
PINK = "E496B3"
BLUE = "B6D6EB"
MUTED = "666862"
BORDER = "D8D1C5"
GREEN = "DDEAD7"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER),
    right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER),
    bottom=Side(style="thin", color=BORDER),
)

ORDER_HEADERS = [
    "Order ID",
    "Created",
    "Fulfilment",
    "Dispatched At",
    "Customer",
    "Email",
    "Merchandise",
    "Shipping Charged",
    "Total Revenue",
    "Stripe Fee",
    "Actual Postage",
    "Packaging Cost",
    "Other Cost",
    "Item Cost",
    "Net Profit",
    "Shipping Service",
    "Tracking Number",
    "Notes",
]

ITEM_HEADERS = [
    "Order ID",
    "Line",
    "Created",
    "Product Key",
    "Product",
    "Qty",
    "Configuration",
    "Lid",
    "Base",
    "Left Button",
    "Right Button",
    "Unit Price",
    "Line Revenue",
    "Unit Cost",
    "Line Cost",
]

PRODUCT_HEADERS = ["Product Key", "Product", "Sale Price", "Unit Cost", "Notes"]


def ledger_path(mode: str = "test") -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    safe_mode = "live" if str(mode).lower() == "live" else "sandbox"
    return DATA_DIR / f"nicehatthanks-orders-{safe_mode}.xlsx"


def pounds(pence: Any) -> float:
    try:
        return round(int(pence) / 100.0, 2)
    except (TypeError, ValueError):
        return 0.0


def order_created(timestamp: Any) -> datetime | None:
    try:
        return datetime.fromtimestamp(int(timestamp))
    except (TypeError, ValueError, OSError):
        return None


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(color=CREAM, bold=True)
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def style_data_area(ws, max_col: int) -> None:
    for row in ws.iter_rows(min_row=2, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")


def set_widths(ws, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def ensure_workbook(path: Path):
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    wb.create_sheet("Summary")
    wb.create_sheet("Orders")
    wb.create_sheet("Items")
    wb.create_sheet("Products")
    return wb


def ensure_sheet(wb, name: str):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(name)


def sync_products(ws) -> None:
    if ws.max_row < 1 or ws.cell(1, 1).value != PRODUCT_HEADERS[0]:
        ws.delete_rows(1, ws.max_row)
        ws.append(PRODUCT_HEADERS)

    existing: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, 1).value
        if key:
            existing[str(key)] = row

    for key, (name, sale_price) in PRODUCT_DEFAULTS.items():
        row = existing.get(key)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row, 1, key)
            ws.cell(row, 4, None)
            ws.cell(row, 5, None)
        ws.cell(row, 2, name)
        ws.cell(row, 3, sale_price)

    style_header(ws)
    style_data_area(ws, len(PRODUCT_HEADERS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(ws.max_row, 1)}"
    set_widths(ws, {1: 22, 2: 30, 3: 13, 4: 13, 5: 34})
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 3).number_format = "£0.00"
        ws.cell(row, 4).number_format = "£0.00"
        ws.cell(row, 4).fill = PatternFill("solid", fgColor="FFF4C7")


def rebuild_items(ws, orders: list[dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(ITEM_HEADERS)

    for order in sorted(orders, key=lambda value: int(value.get("created") or 0)):
        order_id = str(order.get("id") or "")
        created = order_created(order.get("created"))
        for line_number, item in enumerate(order.get("items") or [], start=1):
            colours = item.get("colours") or {}
            row = ws.max_row + 1
            quantity = int(item.get("quantity") or 0)
            unit_price = pounds(item.get("unitAmount"))
            ws.append([
                order_id,
                line_number,
                created,
                item.get("product") or "",
                item.get("name") or item.get("product") or "",
                quantity,
                item.get("configuration") or "",
                colours.get("lid") or "",
                colours.get("base") or "",
                colours.get("leftButton") or "",
                colours.get("rightButton") or "",
                unit_price,
                round(unit_price * quantity, 2),
                f'=IFERROR(VLOOKUP(D{row},Products!$A$2:$D$100,4,FALSE),0)',
                f'=F{row}*N{row}',
            ])

    style_header(ws)
    style_data_area(ws, len(ITEM_HEADERS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{max(ws.max_row, 1)}"
    set_widths(ws, {
        1: 28, 2: 8, 3: 19, 4: 22, 5: 30, 6: 8, 7: 18,
        8: 14, 9: 14, 10: 14, 11: 14, 12: 12, 13: 13, 14: 12, 15: 12,
    })
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 3).number_format = "dd mmm yyyy hh:mm"
        for column in (12, 13, 14, 15):
            ws.cell(row, column).number_format = "£0.00"


def sync_orders(ws, orders: list[dict[str, Any]]) -> None:
    if ws.max_row < 1 or ws.cell(1, 1).value != ORDER_HEADERS[0]:
        ws.delete_rows(1, ws.max_row)
        ws.append(ORDER_HEADERS)

    existing: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        order_id = ws.cell(row, 1).value
        if order_id:
            existing[str(order_id)] = row

    for order in sorted(orders, key=lambda value: int(value.get("created") or 0)):
        order_id = str(order.get("id") or "")
        if not order_id:
            continue
        row = existing.get(order_id)
        if row is None:
            row = ws.max_row + 1
            existing[order_id] = row

        customer = order.get("customer") or {}
        shipping = order.get("shipping") or {}
        fulfilment = order.get("fulfilment") or {}
        total_revenue = pounds(order.get("amountTotal"))
        shipping_charged = pounds(shipping.get("amount"))
        merchandise = round(total_revenue - shipping_charged, 2)

        automatic_values = {
            1: order_id,
            2: order_created(order.get("created")),
            3: fulfilment.get("status") or "pending",
            4: fulfilment.get("dispatchedAt") or "",
            5: customer.get("name") or "",
            6: customer.get("email") or "",
            7: merchandise,
            8: shipping_charged,
            9: total_revenue,
            16: shipping.get("service") or "",
            17: fulfilment.get("trackingNumber") or "",
        }
        for column, value in automatic_values.items():
            ws.cell(row, column, value)

        # J:M and R are deliberately left untouched: these are the owner's
        # manual accounting fields and must survive automatic refreshes.
        ws.cell(row, 14, f'=SUMIF(Items!$A:$A,A{row},Items!$O:$O)')
        ws.cell(row, 15, f'=I{row}-SUM(J{row}:N{row})')

    style_header(ws)
    style_data_area(ws, len(ORDER_HEADERS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{max(ws.max_row, 1)}"
    set_widths(ws, {
        1: 30, 2: 19, 3: 13, 4: 25, 5: 24, 6: 32,
        7: 13, 8: 15, 9: 14, 10: 12, 11: 14, 12: 14,
        13: 12, 14: 12, 15: 13, 16: 38, 17: 24, 18: 36,
    })

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 2).number_format = "dd mmm yyyy hh:mm"
        for column in range(7, 16):
            ws.cell(row, column).number_format = "£0.00"
        for column in (10, 11, 12, 13, 18):
            ws.cell(row, column).fill = PatternFill("solid", fgColor="FFF4C7")
        status = str(ws.cell(row, 3).value or "").lower()
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=GREEN if status == "dispatched" else "F4D7E2")


def rebuild_summary(ws, mode: str) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "NiceHatThanks Finance"
    ws["A1"].font = Font(size=20, bold=True, color=DARK)
    ws["A2"] = "LIVE ledger" if str(mode).lower() == "live" else "SANDBOX ledger - test orders only"
    ws["A2"].font = Font(color=MUTED, italic=True)

    labels = [
        ("A4", "Paid orders", '=MAX(COUNTA(Orders!A:A)-1,0)'),
        ("A5", "Total revenue", '=SUM(Orders!I:I)'),
        ("A6", "Merchandise revenue", '=SUM(Orders!G:G)'),
        ("A7", "Shipping charged", '=SUM(Orders!H:H)'),
        ("A9", "Stripe fees", '=SUM(Orders!J:J)'),
        ("A10", "Actual postage", '=SUM(Orders!K:K)'),
        ("A11", "Packaging", '=SUM(Orders!L:L)'),
        ("A12", "Other costs", '=SUM(Orders!M:M)'),
        ("A13", "Product costs", '=SUM(Orders!N:N)'),
        ("A15", "Net profit", '=SUM(Orders!O:O)'),
    ]
    for label_cell, label, formula in labels:
        row = ws[label_cell].row
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=CREAM)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=WHITE)
        ws.cell(row, 1).border = THIN_BORDER
        ws.cell(row, 2).border = THIN_BORDER

    ws["B4"].number_format = "0"
    for row in (5, 6, 7, 9, 10, 11, 12, 13, 15):
        ws.cell(row, 2).number_format = "£0.00"

    ws["A18"] = "Cost setup"
    ws["A18"].font = Font(size=12, bold=True, color=DARK)
    ws["A19"] = (
        "Fill Unit Cost on the Products sheet, plus Stripe Fee / Actual Postage / "
        "Packaging Cost / Other Cost on Orders. Yellow cells are manual inputs. "
        "Until those are filled in, Net Profit is incomplete."
    )
    ws.merge_cells("A19:F21")
    ws["A19"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A19"].fill = PatternFill("solid", fgColor="FFF4C7")
    ws["A19"].border = THIN_BORDER

    ws["A23"] = "Last synced"
    ws["B23"] = datetime.now()
    ws["B23"].number_format = "dd mmm yyyy hh:mm:ss"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    for column in "CDEF":
        ws.column_dimensions[column].width = 14


def sync_ledger(orders: list[dict[str, Any]], mode: str = "test") -> Path:
    path = ledger_path(mode)
    wb = ensure_workbook(path)

    summary = ensure_sheet(wb, "Summary")
    orders_ws = ensure_sheet(wb, "Orders")
    items_ws = ensure_sheet(wb, "Items")
    products_ws = ensure_sheet(wb, "Products")

    sync_products(products_ws)
    rebuild_items(items_ws, orders)
    sync_orders(orders_ws, orders)
    rebuild_summary(summary, mode)

    wb.active = wb.sheetnames.index("Summary")
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
